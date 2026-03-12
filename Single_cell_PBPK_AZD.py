#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import numpy as np
import pandas as pd
from scipy.integrate import solve_ivp
import matplotlib.pyplot as plt


DEFAULT_PARAMS = dict(
    TSTOP=24.0,
    POINTS=800,

    CO=368.0,
    Qbo=9.2, Qbr=25.1, Qfa=12.5, Qhe=8.6, Qki=42.2, Qli=15.5,
    Qlu=200.4, Qmu=43.6, Qpv=36.0, Qgi=36.0, Qrb=9.7, Qco=200.4,

    BW=73.0,
    Vbo=10.6, VbrV=0.03, VbrIF=0.01, VbrIC=1.41,
    Vfa=18.6, Vhe=0.5, Vki=0.4, Vli=1.9, Vlu=0.9, Vmu=30.4,
    Vpv=0.46, Vgi=1.6, Vrb=3.5, Vapl=0.22, Vvpl=0.50,

    Rbo=129.0, Rbr1=0.49, Rbr2=9.6,
    Rfa=376.0, Rhe=49.0, Rki=26.0, Rli=4.3,
    Rlu=7.0, Rmu=8.0, Rgi=28.0, Rrb=33.0,

    CLbreffPgp=2.7,
    CLbreffABCG2=2.96,
    PSB=12.6,
    PSB2=54.2,
    CLbrUptake=105.4,
    fup=0.2,

    OralDose=400.0,
    kabs=2.1,
    F=0.7,

    CLliInt=250.0,
    CLkiInt=5.5,
)



def oral_absorption_rate(t, kabs, F, dose_mg):
    return 0.0 if t < 0.0 else (kabs * F * dose_mg * np.exp(-kabs * t))


def loguniform(low, high, size=None, rng=None):
    rng = np.random.default_rng() if rng is None else rng
    low = float(low); high = float(high)
    if low <= 0 or high <= 0 or high <= low:
        raise ValueError("loguniform requires 0 < low < high")
    return np.exp(rng.uniform(np.log(low), np.log(high), size=size))


def sample_clusters(n_cells, n_clusters=4, cluster_fracs=None, rng=None):
    rng = np.random.default_rng() if rng is None else rng
    if cluster_fracs is None:
        cluster_fracs = np.full(n_clusters, 1.0 / n_clusters, dtype=float)
    else:
        cluster_fracs = np.asarray(cluster_fracs, float)
        cluster_fracs = cluster_fracs / cluster_fracs.sum()

    counts = np.floor(cluster_fracs * n_cells).astype(int)
    rem = n_cells - counts.sum()
    if rem > 0:
        frac_parts = (cluster_fracs * n_cells) - np.floor(cluster_fracs * n_cells)
        add_idx = np.argsort(-frac_parts)[:rem]
        counts[add_idx] += 1

    cl = np.concatenate([np.full(counts[k], k, dtype=int) for k in range(n_clusters)])
    rng.shuffle(cl)
    return cl


def nb_multiplier_from_counts(n, mu_mult, theta, mu_count=100.0, rng=None, floor=1e-12):
 
     rng = np.random.default_rng() if rng is None else rng
    mu_mult = float(mu_mult)
    theta = float(theta)
    mu_count = float(mu_count)

    if mu_mult <= 0:
        return np.full(n, floor, dtype=float)
    if theta <= 0:
        raise ValueError("theta must be > 0")
    if mu_count <= 0:
        raise ValueError("mu_count must be > 0")

    lam = rng.gamma(shape=theta, scale=mu_count/theta, size=n)
    k = rng.poisson(lam).astype(float)
    k = np.maximum(k, 1.0)  # strictly positive
    w = (k / max(k.mean(), floor)) * mu_mult
    return np.maximum(w, floor)


def sample_clustered_triplet_multipliers_optionA(
    n_cells,
    cluster_id,
    uptake_mus, uptake_thetas,
    pgp_mus,    pgp_thetas,
    abcg2_mus,  abcg2_thetas,
    mu_count=100.0,
    seed=10,
    floor=1e-12
):

    cid = np.asarray(cluster_id, dtype=int)
    K = int(cid.max()) + 1
    uptake_mus = np.asarray(uptake_mus, float); uptake_thetas = np.asarray(uptake_thetas, float)
    pgp_mus    = np.asarray(pgp_mus, float);    pgp_thetas    = np.asarray(pgp_thetas, float)
    abcg2_mus  = np.asarray(abcg2_mus, float);  abcg2_thetas  = np.asarray(abcg2_thetas, float)
    if uptake_mus.shape != (K,) or uptake_thetas.shape != (K,):
        raise ValueError("uptake mus/thetas must be (K,)")
    if pgp_mus.shape != (K,) or pgp_thetas.shape != (K,):
        raise ValueError("pgp mus/thetas must be (K,)")
    if abcg2_mus.shape != (K,) or abcg2_thetas.shape != (K,):
        raise ValueError("abcg2 mus/thetas must be (K,)")

    uptake_W = np.empty(n_cells, float)
    pgp_W    = np.empty(n_cells, float)
    abcg2_W  = np.empty(n_cells, float)

    for k in range(K):
        idx = np.where(cid == k)[0]
        if idx.size == 0:
            continue

        rng_u = np.random.default_rng(seed + 10007*k + 1)
        rng_p = np.random.default_rng(seed + 10007*k + 2)
        rng_a = np.random.default_rng(seed + 10007*k + 3)

        uptake_W[idx] = nb_multiplier_from_counts(idx.size, uptake_mus[k], uptake_thetas[k], mu_count=mu_count, rng=rng_u, floor=floor)
        pgp_W[idx]    = nb_multiplier_from_counts(idx.size, pgp_mus[k],    pgp_thetas[k],    mu_count=mu_count, rng=rng_p, floor=floor)
        abcg2_W[idx]  = nb_multiplier_from_counts(idx.size, abcg2_mus[k],  abcg2_thetas[k],  mu_count=mu_count, rng=rng_a, floor=floor)

    return uptake_W, pgp_W, abcg2_W


def random_cluster_parameters(n_clusters, mean_range, theta_range, seed=None):
 
    rng = np.random.default_rng(seed)
    means = rng.uniform(mean_range[0], mean_range[1], size=n_clusters)
    thetas = loguniform(theta_range[0], theta_range[1], size=n_clusters, rng=rng)
    return means.tolist(), thetas.tolist()


# ============================================================
# Allocate organ total clearance to explicit cells + bulk (conserving organ total)
# ============================================================
def alloc_total_to_cells_and_bulk(total_CL, multipliers, n_het, N_tot, name="CL"):

    total_CL = float(total_CL)
    CL_per_cell_avg = total_CL / int(N_tot)

    if multipliers is None:
        mult = np.ones(int(n_het), dtype=float)
    else:
        mult = np.asarray(multipliers, dtype=float)
        if mult.shape != (int(n_het),):
            raise ValueError(f"{name}_multipliers must have shape ({n_het},)")
        mult = np.maximum(mult, 1e-12)

    mult = mult / np.mean(mult)  # conserve total
    CL_i = CL_per_cell_avg * mult
    CL_explicit_total = float(np.sum(CL_i))
    CL_bulk_total = max(total_CL - CL_explicit_total, 0.0)
    return CL_i, CL_bulk_total


# ============================================================
# RHS ODE
# ============================================================
def rhs_azd1775_multicell_brain(
    t, y, p,
    VbrIF_i, V_IF_bulk,
    VbrIC_i, V_IC_bulk,
    PSB_per_cell, PSB_bulk_total,
    PSB2_per_cell, PSB2_bulk_total,
    CLbrUptake_i, CLbrUptake_bulk_total,
    CLpgp_i, CLpgp_bulk_total,
    CLabcg2_i, CLabcg2_bulk_total,
):
    n_het = len(VbrIC_i)

    i_Abo = 0
    i_AbrV = 1
    i_AbrIF_s = 2
    i_AbrIF_e = i_AbrIF_s + n_het
    i_AIF_bulk = i_AbrIF_e
    i_AbrIC_s = i_AIF_bulk + 1
    i_AbrIC_e = i_AbrIC_s + n_het
    i_AIC_bulk = i_AbrIC_e

    i_Afa  = i_AIC_bulk + 1
    i_Ahe  = i_AIC_bulk + 2
    i_Aki  = i_AIC_bulk + 3
    i_Ali  = i_AIC_bulk + 4
    i_Alu  = i_AIC_bulk + 5
    i_Amu  = i_AIC_bulk + 6
    i_Arb  = i_AIC_bulk + 7
    i_Agi  = i_AIC_bulk + 8
    i_Apv  = i_AIC_bulk + 9
    i_Avpl = i_AIC_bulk + 10
    i_Aapl = i_AIC_bulk + 11

    i_AUCbr  = i_AIC_bulk + 12
    i_AUCli  = i_AIC_bulk + 13
    i_AUCvpl = i_AIC_bulk + 14

    Abo = y[i_Abo]
    AbrV = y[i_AbrV]
    AbrIF_i = y[i_AbrIF_s:i_AbrIF_e]
    AIF_bulk = y[i_AIF_bulk]
    AbrIC_i = y[i_AbrIC_s:i_AbrIC_e]
    AIC_bulk = y[i_AIC_bulk]

    Afa  = y[i_Afa]
    Ahe  = y[i_Ahe]
    Aki  = y[i_Aki]
    Ali  = y[i_Ali]
    Alu  = y[i_Alu]
    Amu  = y[i_Amu]
    Arb  = y[i_Arb]
    Agi  = y[i_Agi]
    Apv  = y[i_Apv]
    Avpl = y[i_Avpl]
    Aapl = y[i_Aapl]

    # params
    Qbo, Qbr, Qfa, Qhe, Qki, Qli, Qmu, Qpv, Qgi, Qrb, Qco = (
        p["Qbo"], p["Qbr"], p["Qfa"], p["Qhe"], p["Qki"], p["Qli"], p["Qmu"],
        p["Qpv"], p["Qgi"], p["Qrb"], p["Qco"]
    )
    Vbo, VbrV, Vfa, Vhe, Vki, Vli, Vlu, Vmu, Vgi, Vpv, Vrb, Vapl, Vvpl = (
        p["Vbo"], p["VbrV"], p["Vfa"], p["Vhe"], p["Vki"],
        p["Vli"], p["Vlu"], p["Vmu"], p["Vgi"], p["Vpv"], p["Vrb"], p["Vapl"], p["Vvpl"]
    )
    Rbo, Rbr1, Rbr2, Rfa, Rhe, Rki, Rli, Rlu, Rmu, Rgi, Rrb = (
        p["Rbo"], p["Rbr1"], p["Rbr2"], p["Rfa"], p["Rhe"], p["Rki"], p["Rli"],
        p["Rlu"], p["Rmu"], p["Rgi"], p["Rrb"]
    )
    kabs, F, OralDose = p["kabs"], p["F"], p["OralDose"]
    CLliInt, CLkiInt, fup = p["CLliInt"], p["CLkiInt"], p["fup"]

    # concentrations
    Capl = Aapl / Vapl
    Cvpl = Avpl / Vvpl

    Cbo = Abo / Vbo
    Cfa = Afa / Vfa
    Che = Ahe / Vhe
    Cki = Aki / Vki
    Cli = Ali / Vli
    Clu = Alu / Vlu
    Cmu = Amu / Vmu
    Crb = Arb / Vrb
    Cgi = Agi / Vgi
    Cpv = Apv / Vpv

    CbrV = AbrV / VbrV
    CbrIF_i = AbrIF_i / VbrIF_i
    CbrIF_bulk = AIF_bulk / V_IF_bulk
    CbrIC_i = AbrIC_i / VbrIC_i
    CbrIC_bulk = AIC_bulk / V_IC_bulk

    # Bone
    rboIn = Qbo * Capl
    rboOut = Qbo * Cbo / Rbo
    dAbo = rboIn - rboOut

    # Brain vascular perfusion
    rbrIn = Qbr * Capl
    rbrOut = Qbr * CbrV

    # V <-> IF (split PSB)
    tVIF_i = PSB_per_cell * (CbrV - CbrIF_i / Rbr1)
    tVIF_bulk = PSB_bulk_total * (CbrV - CbrIF_bulk / Rbr1)

    # Uptake V -> IF (heterogeneous)
    tUp_i = CLbrUptake_i * CbrV * fup
    tUp_bulk = CLbrUptake_bulk_total * CbrV * fup

    # Efflux IF -> V (heterogeneous)
    tPgp_i = CLpgp_i * (CbrIF_i / Rbr1)
    tPgp_bulk = CLpgp_bulk_total * (CbrIF_bulk / Rbr1)

    tAbc_i = CLabcg2_i * (CbrIF_i / Rbr1)
    tAbc_bulk = CLabcg2_bulk_total * (CbrIF_bulk / Rbr1)

    # IF <-> IC diffusion (split PSB2)
    tIFIC_i = PSB2_per_cell * (CbrIF_i - CbrIC_i / Rbr2)
    tIFIC_bulk = PSB2_bulk_total * (CbrIF_bulk - CbrIC_bulk / Rbr2)

    dAbrV = (
        rbrIn - rbrOut
        - np.sum(tVIF_i) - tVIF_bulk
        - np.sum(tUp_i)  - tUp_bulk
        + np.sum(tPgp_i) + tPgp_bulk
        + np.sum(tAbc_i) + tAbc_bulk
    )
    dAbrIF_i = tVIF_i + tUp_i - tPgp_i - tAbc_i - tIFIC_i
    dAIF_bulk = tVIF_bulk + tUp_bulk - tPgp_bulk - tAbc_bulk - tIFIC_bulk
    dAbrIC_i = tIFIC_i
    dAIC_bulk = tIFIC_bulk

    # AUC brain (total brain concentration)
    AbrIF_total = np.sum(AbrIF_i) + AIF_bulk
    AbrIC_total = np.sum(AbrIC_i) + AIC_bulk
    VbrIF_total = np.sum(VbrIF_i) + V_IF_bulk
    VbrIC_total = np.sum(VbrIC_i) + V_IC_bulk
    Cbr_total = (AbrV + AbrIF_total + AbrIC_total) / (VbrV + VbrIF_total + VbrIC_total)
    dAUCbr = Cbr_total

    # Fat
    dAfa = Qfa * Capl - Qfa * Cfa / Rfa

    # Heart
    dAhe = Qhe * Capl - Qhe * Che / Rhe

    # Kidney
    rkiIn = Qki * Capl
    rkiOut = Qki * Cki / Rki
    ekiCLint = CLkiInt * Cki * (fup / Rki)
    dAki = rkiIn - rkiOut - ekiCLint

    # Liver
    rliIn = Qli * Capl + Qpv * Cpv
    rliOut = (Qli + Qpv) * Cli / Rli
    eliCLint = CLliInt * Cli * (fup / Rli)
    dAli = rliIn - rliOut - eliCLint
    dAUCli = Cli

    # Lung
    rvplOut = Qco * Cvpl
    rluIn = rvplOut
    rluOut = Qco * Clu / Rlu
    dAlu = rluIn - rluOut

    # Muscle
    dAmu = Qmu * Capl - Qmu * Cmu / Rmu

    # Rest of body
    dArb = Qrb * Capl - Qrb * Crb / Rrb

    # GI + absorption
    rgiIn = Qgi * Capl
    rgiOut = Qgi * Cgi / Rgi
    rAbsgi = oral_absorption_rate(t, kabs=kabs, F=F, dose_mg=OralDose)
    dAgi = rgiIn - rgiOut + rAbsgi

    # Portal vein
    dApv = rgiOut - Qpv * Cpv

    # Venous plasma
    rvplIn = (rboOut + rbrOut + Qfa * Cfa / Rfa + Qhe * Che / Rhe + rkiOut + rliOut + Qmu * Cmu / Rmu + Qrb * Crb / Rrb)
    dAvpl = rvplIn - rvplOut
    dAUCvpl = Cvpl

    # Arterial plasma
    dAapl = rluOut - Qco * Capl

    dydt = np.zeros_like(y)
    dydt[i_Abo] = dAbo
    dydt[i_AbrV] = dAbrV
    dydt[i_AbrIF_s:i_AbrIF_e] = dAbrIF_i
    dydt[i_AIF_bulk] = dAIF_bulk
    dydt[i_AbrIC_s:i_AbrIC_e] = dAbrIC_i
    dydt[i_AIC_bulk] = dAIC_bulk

    dydt[i_Afa] = dAfa
    dydt[i_Ahe] = dAhe
    dydt[i_Aki] = dAki
    dydt[i_Ali] = dAli
    dydt[i_Alu] = dAlu
    dydt[i_Amu] = dAmu
    dydt[i_Arb] = dArb
    dydt[i_Agi] = dAgi
    dydt[i_Apv] = dApv
    dydt[i_Avpl] = dAvpl
    dydt[i_Aapl] = dAapl

    dydt[i_AUCbr] = dAUCbr
    dydt[i_AUCli] = dAUCli
    dydt[i_AUCvpl] = dAUCvpl
    return dydt


# ============================================================
# Simulation wrapper
# ============================================================
def simulate_azd1775_multicell_brain_with_bulk(
    params=None,
    n_het_cells=2000,
    N_total_cells=3.75e8,
    uptake_multipliers=None,
    pgp_multipliers=None,
    abcg2_multipliers=None,
    t_eval=None,
    debug=True
):
    p = DEFAULT_PARAMS.copy()
    if params is not None:
        p.update(params)

    TSTOP = float(p["TSTOP"])
    POINTS = int(p["POINTS"])
    if t_eval is None:
        t_eval = np.linspace(0.0, TSTOP, POINTS + 1)

    n_het = int(n_het_cells)
    N_tot = int(N_total_cells)
    if N_tot <= n_het:
        raise ValueError("N_total_cells must be > n_het_cells to have a bulk pool.")
    N_bulk = N_tot - n_het

    # volumes split across cells
    VbrIF_total = float(p["VbrIF"])
    VbrIC_total = float(p["VbrIC"])
    v_if_cell = VbrIF_total / N_tot
    v_ic_cell = VbrIC_total / N_tot

    VbrIF_i = np.full(n_het, v_if_cell, dtype=float)
    V_IF_bulk = v_if_cell * N_bulk
    VbrIC_i = np.full(n_het, v_ic_cell, dtype=float)
    V_IC_bulk = v_ic_cell * N_bulk

    # barrier splits
    PSB_total = float(p["PSB"])
    PSB_per_cell = PSB_total / N_tot
    PSB_bulk_total = max(PSB_total - PSB_per_cell * n_het, 0.0)

    PSB2_total = float(p["PSB2"])
    PSB2_per_cell = PSB2_total / N_tot
    PSB2_bulk_total = max(PSB2_total - PSB2_per_cell * n_het, 0.0)

    # allocate heterogeneous CLs (conserve totals)
    CL_upt_i, CL_upt_bulk = alloc_total_to_cells_and_bulk(
        float(p["CLbrUptake"]), uptake_multipliers, n_het, N_tot, name="uptake"
    )
    CL_pgp_i, CL_pgp_bulk = alloc_total_to_cells_and_bulk(
        float(p["CLbreffPgp"]), pgp_multipliers, n_het, N_tot, name="pgp"
    )
    CL_abc_i, CL_abc_bulk = alloc_total_to_cells_and_bulk(
        float(p["CLbreffABCG2"]), abcg2_multipliers, n_het, N_tot, name="abcg2"
    )

    if debug:
        print("\n--- AZD multicell brain setup (Option A; with bulk pool) ---")
        print(f"n_het_cells   = {n_het}")
        print(f"N_total_cells = {N_tot}")
        print(f"N_bulk        = {N_bulk}")
        print(f"VbrIF_total   = {VbrIF_total:.6f} L, v_if_cell={v_if_cell:.3e} L")
        print(f"VbrIC_total   = {VbrIC_total:.6f} L, v_ic_cell={v_ic_cell:.3e} L")
        print(f"Sum(VbrIF_i)  = {VbrIF_i.sum():.6f} L, V_IF_bulk={V_IF_bulk:.6f} L")
        print(f"Sum(VbrIC_i)  = {VbrIC_i.sum():.6f} L, V_IC_bulk={V_IC_bulk:.6f} L")
        print(f"PSB_total={PSB_total:.3f}, PSB_per_cell={PSB_per_cell:.3e}, PSB_bulk={PSB_bulk_total:.3f}")
        print(f"PSB2_total={PSB2_total:.3f}, PSB2_per_cell={PSB2_per_cell:.3e}, PSB2_bulk={PSB2_bulk_total:.3f}")
        print(f"CLbrUptake total={p['CLbrUptake']:.3f}, explicit={CL_upt_i.sum():.3f}, bulk={CL_upt_bulk:.3f}")
        print(f"CLbreffPgp total={p['CLbreffPgp']:.3f}, explicit={CL_pgp_i.sum():.3f}, bulk={CL_pgp_bulk:.3f}")
        print(f"CLbreffABCG2 total={p['CLbreffABCG2']:.3f}, explicit={CL_abc_i.sum():.3f}, bulk={CL_abc_bulk:.3f}")
        print("----------------------------------------------------------")

    n_states = (15 + 3) + (2 * n_het) + 2
    y0 = np.zeros(n_states, dtype=float)

    sol = solve_ivp(
        fun=lambda t, y: rhs_azd1775_multicell_brain(
            t, y, p,
            VbrIF_i, V_IF_bulk,
            VbrIC_i, V_IC_bulk,
            PSB_per_cell, PSB_bulk_total,
            PSB2_per_cell, PSB2_bulk_total,
            CL_upt_i, CL_upt_bulk,
            CL_pgp_i, CL_pgp_bulk,
            CL_abc_i, CL_abc_bulk,
        ),
        t_span=(t_eval[0], t_eval[-1]),
        y0=y0,
        t_eval=t_eval,
        method="BDF",
        rtol=1e-7,
        atol=1e-10
    )

    # unpack indices
    i_Abo = 0
    i_AbrV = 1
    i_AbrIF_s = 2
    i_AbrIF_e = i_AbrIF_s + n_het
    i_AIF_bulk = i_AbrIF_e
    i_AbrIC_s = i_AIF_bulk + 1
    i_AbrIC_e = i_AbrIC_s + n_het
    i_AIC_bulk = i_AbrIC_e

    i_Aki = i_AIC_bulk + 3
    i_Ali = i_AIC_bulk + 4
    i_Avpl = i_AIC_bulk + 10
    i_Aapl = i_AIC_bulk + 11
    i_AUCbr = i_AIC_bulk + 12
    i_AUCli = i_AIC_bulk + 13
    i_AUCvpl = i_AIC_bulk + 14

    AbrV = sol.y[i_AbrV, :]
    AbrIF_mat_amt = sol.y[i_AbrIF_s:i_AbrIF_e, :]
    AIF_bulk_amt = sol.y[i_AIF_bulk, :]
    AbrIC_mat_amt = sol.y[i_AbrIC_s:i_AbrIC_e, :]
    AIC_bulk_amt = sol.y[i_AIC_bulk, :]

    Aki = sol.y[i_Aki, :]
    Ali = sol.y[i_Ali, :]
    Avpl = sol.y[i_Avpl, :]
    Aapl = sol.y[i_Aapl, :]

    t = sol.t
    Cvpl = Avpl / p["Vvpl"]
    CbrV = AbrV / p["VbrV"]

    CbrIF_mat = AbrIF_mat_amt / VbrIF_i[:, None]
    CbrIC_mat = AbrIC_mat_amt / VbrIC_i[:, None]

    # global IF/IC and total brain
    AbrIF_total = AbrIF_mat_amt.sum(axis=0) + AIF_bulk_amt
    AbrIC_total = AbrIC_mat_amt.sum(axis=0) + AIC_bulk_amt
    CbrIF_global = AbrIF_total / (VbrIF_i.sum() + V_IF_bulk)
    CbrIC_global = AbrIC_total / (VbrIC_i.sum() + V_IC_bulk)
    Cbr_total = (AbrV + AbrIF_total + AbrIC_total) / (p["VbrV"] + (VbrIF_i.sum() + V_IF_bulk) + (VbrIC_i.sum() + V_IC_bulk))

    df = pd.DataFrame({
        "Time_hr": t,
        "Cvpl_mg_per_L": Cvpl,
        "CbrV_mg_per_L": CbrV,
        "CbrIF_global_mg_per_L": CbrIF_global,
        "CbrIC_global_mg_per_L": CbrIC_global,
        "Cbr_total_mg_per_L": Cbr_total,
        "Cli_mg_per_L": Ali / p["Vli"],
        "Cki_mg_per_L": Aki / p["Vki"],
        "AUCbr": sol.y[i_AUCbr, :],
        "AUCli": sol.y[i_AUCli, :],
        "AUCvpl": sol.y[i_AUCvpl, :],
    })

    return sol, df, CbrIF_mat, CbrIC_mat


# ============================================================
# Plot helpers
# ============================================================
def plot_clusters_2x2_bands(t, mat, cl, title_prefix, q_lo=0.05, q_hi=0.95, sharey=True):
    t = np.asarray(t)
    mat = np.asarray(mat)
    cl = np.asarray(cl, dtype=int)

    K = int(cl.max()) + 1
    Kplot = min(K, 4)
    if K != 4:
        print(f"Warning: K={K}. This function is designed for K=4 (2x2). Plotting first {Kplot} clusters.")

    fig, axes = plt.subplots(2, 2, figsize=(12, 8), sharex=True, sharey=sharey)
    axes = axes.ravel()

    for k in range(Kplot):
        ax = axes[k]
        idx = np.where(cl == k)[0]
        if idx.size == 0:
            ax.set_title(f"Cluster {k} (n=0)")
            ax.axis("off")
            continue

        X = mat[idx, :]                   # (n_k, T)
        lo = np.quantile(X, q_lo, axis=0)
        hi = np.quantile(X, q_hi, axis=0)
        mu = np.mean(X, axis=0)

        ax.fill_between(t, lo, hi, alpha=0.25)
        ax.plot(t, mu, lw=2.5)
        ax.set_title(f"Cluster {k} (n={idx.size})")
        ax.set_xlabel("Time (hr)")
        ax.set_ylabel("mg/L")
        ax.grid(True, ls="--", alpha=0.5)

    for j in range(Kplot, 4):
        axes[j].axis("off")

    fig.suptitle(f"{title_prefix}: mean + {int(q_lo*100)}–{int(q_hi*100)}% band", y=1.02)
    plt.tight_layout()
    plt.show()


# ============================================================
# Excel export (same structure you used, minor cleanup)
# ============================================================
def export_azd_excel(
    out_xlsx,
    sol, df,
    CbrIF_mat, CbrIC_mat,
    cl, N_tot,
    n_keep_per_cluster=2500,
    seed_save=123
):
    import openpyxl  # ensures engine is available

    t = df["Time_hr"].to_numpy()
    cl = np.asarray(cl, dtype=int)
    n_het = int(cl.size)
    n_clusters = int(cl.max()) + 1

    CbrIF_mat = np.asarray(CbrIF_mat)
    CbrIC_mat = np.asarray(CbrIC_mat)
    assert CbrIF_mat.shape == (n_het, t.size)
    assert CbrIC_mat.shape == (n_het, t.size)

    # indices (must match simulate() indexing)
    i_Abo = 0
    i_AbrV = 1
    i_AbrIF_s = 2
    i_AbrIF_e = i_AbrIF_s + n_het
    i_AIF_bulk = i_AbrIF_e
    i_AbrIC_s = i_AIF_bulk + 1
    i_AbrIC_e = i_AbrIC_s + n_het
    i_AIC_bulk = i_AbrIC_e

    AIF_bulk_amt = sol.y[i_AIF_bulk, :]
    AIC_bulk_amt = sol.y[i_AIC_bulk, :]

    # volumes (same splitting as in simulate_azd1775_multicell_brain_with_bulk)
    p = DEFAULT_PARAMS.copy()  # this matches your run if you didn't override volumes
    N_tot_int = int(N_tot)
    N_bulk_int = N_tot_int - n_het
    v_if_cell = float(p["VbrIF"]) / N_tot_int
    v_ic_cell = float(p["VbrIC"]) / N_tot_int
    V_IF_bulk = v_if_cell * N_bulk_int
    V_IC_bulk = v_ic_cell * N_bulk_int

    CbrIF_bulk = AIF_bulk_amt / V_IF_bulk
    CbrIC_bulk = AIC_bulk_amt / V_IC_bulk

    # choose subset of cells per cluster
    rng = np.random.default_rng(seed_save)
    kept_idx = []
    for k in range(n_clusters):
        idx_k = np.where(cl == k)[0]
        if idx_k.size == 0:
            continue
        take = min(n_keep_per_cluster, idx_k.size)
        kept_idx.append(rng.choice(idx_k, size=take, replace=False))
    kept_idx = np.concatenate(kept_idx) if len(kept_idx) else np.array([], dtype=int)
    kept_idx = kept_idx[np.lexsort((kept_idx, cl[kept_idx]))]

    print(f"Saving {kept_idx.size} cells total ({n_keep_per_cluster} per cluster max, n_clusters={n_clusters}).")
    print(f"Excel output: {out_xlsx}")

    venous_df = pd.DataFrame({"Time_hr": t, "Cvpl_mg_per_L": df["Cvpl_mg_per_L"].to_numpy()})
    vascular_df = pd.DataFrame({"Time_hr": t, "CbrV_mg_per_L": df["CbrV_mg_per_L"].to_numpy()})

    if_df = pd.DataFrame({"Time_hr": t})
    for i in kept_idx:
        k = int(cl[i])
        if_df[f"IF_cluster{k}_cell{i}"] = CbrIF_mat[i, :]

    ic_df = pd.DataFrame({"Time_hr": t})
    for i in kept_idx:
        k = int(cl[i])
        ic_df[f"IC_cluster{k}_cell{i}"] = CbrIC_mat[i, :]

    if_bulk_df = pd.DataFrame({"Time_hr": t, "CbrIF_bulk_mg_per_L": CbrIF_bulk})
    ic_bulk_df = pd.DataFrame({"Time_hr": t, "CbrIC_bulk_mg_per_L": CbrIC_bulk})

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        venous_df.to_excel(writer, sheet_name="EQ_S1_venous_plasma", index=False)
        vascular_df.to_excel(writer, sheet_name="EQ_S9_brain_vascular", index=False)

        # interstitial: row 2 contains cluster labels for each cell col, data starts row 3
        if_df.to_excel(writer, sheet_name="EQ_S10_brain_interstitial", index=False, startrow=2)
        ws_if = writer.sheets["EQ_S10_brain_interstitial"]
        for col_idx, col_name in enumerate(if_df.columns, start=1):
            if col_name.startswith("IF_cluster") and "_cell" in col_name:
                k = int(col_name.split("IF_cluster", 1)[1].split("_cell", 1)[0])
                ws_if.cell(row=2, column=col_idx, value=f"cluster_{k}")
            else:
                ws_if.cell(row=2, column=col_idx, value="")

        if_bulk_df.to_excel(writer, sheet_name="EQ_S11_brain_interstitial_bulk", index=False)

        ic_df.to_excel(writer, sheet_name="EQ_S12_brain_intracellular", index=False, startrow=2)
        ws_ic = writer.sheets["EQ_S12_brain_intracellular"]
        for col_idx, col_name in enumerate(ic_df.columns, start=1):
            if col_name.startswith("IC_cluster") and "_cell" in col_name:
                k = int(col_name.split("IC_cluster", 1)[1].split("_cell", 1)[0])
                ws_ic.cell(row=2, column=col_idx, value=f"cluster_{k}")
            else:
                ws_ic.cell(row=2, column=col_idx, value="")

        ic_bulk_df.to_excel(writer, sheet_name="EQ_S13_brain_intracellular_bulk", index=False)

    print("Done.")



if __name__ == "__main__":


    params = DEFAULT_PARAMS.copy()

    n_clusters = 4
    n_het = 10000
    N_tot = 450_000_000

    MU_COUNT = 100.0

    # Cluster fractions
    cluster_fracs = [0.25, 0.25, 0.25, 0.25]

    # Means ranges (same as your previous AZD script)
    UPTAKE_MEAN_RANGE = (0.6, 1.2)
    PGP_MEAN_RANGE    = (0.5, 2.5)
    ABCG2_MEAN_RANGE  = (0.5, 2.5)

    # Theta range (log-uniform)
    THETA_RANGE = (1.0, 20.0)

    # Sample per-cluster parameters (keep fixed seeds for reproducibility)
    uptake_mus, uptake_thetas = random_cluster_parameters(n_clusters, UPTAKE_MEAN_RANGE, THETA_RANGE, seed=101)
    pgp_mus,    pgp_thetas    = random_cluster_parameters(n_clusters, PGP_MEAN_RANGE,    THETA_RANGE, seed=202)
    abcg2_mus,  abcg2_thetas  = random_cluster_parameters(n_clusters, ABCG2_MEAN_RANGE,  THETA_RANGE, seed=303)

    print("\n--- AZD Option A sampling ---")
    print("MU_COUNT:", MU_COUNT)
    print("uptake mus:  ", uptake_mus)
    print("uptake theta:", uptake_thetas)
    print("pgp mus:     ", pgp_mus)
    print("pgp theta:   ", pgp_thetas)
    print("abcg2 mus:   ", abcg2_mus)
    print("abcg2 theta: ", abcg2_thetas)

    # Shared cluster labels
    cl = sample_clusters(n_het, n_clusters=n_clusters, cluster_fracs=cluster_fracs, rng=np.random.default_rng(123))

    # Sample clustered multipliers (Option A)
    uptake_W, pgp_W, abcg2_W = sample_clustered_triplet_multipliers_optionA(
        n_cells=n_het,
        cluster_id=cl,
        uptake_mus=uptake_mus, uptake_thetas=uptake_thetas,
        pgp_mus=pgp_mus,       pgp_thetas=pgp_thetas,
        abcg2_mus=abcg2_mus,   abcg2_thetas=abcg2_thetas,
        mu_count=MU_COUNT,
        seed=10
    )

    # Sanity: summarize multipliers by cluster
    print("\nCluster summary (multipliers):")
    for k in range(n_clusters):
        idx = np.where(cl == k)[0]
        print(f"  cluster {k}: n={idx.size}")
        print(f"    uptake mean/std: {uptake_W[idx].mean():.3f} / {uptake_W[idx].std():.3f}")
        print(f"    pgp    mean/std: {pgp_W[idx].mean():.3f} / {pgp_W[idx].std():.3f}")
        print(f"    abcg2  mean/std: {abcg2_W[idx].mean():.3f} / {abcg2_W[idx].std():.3f}")

    # Run simulation
    sol, df, CbrIF_mat, CbrIC_mat = simulate_azd1775_multicell_brain_with_bulk(
        params=params,
        n_het_cells=n_het,
        N_total_cells=N_tot,
        uptake_multipliers=uptake_W,
        pgp_multipliers=pgp_W,
        abcg2_multipliers=abcg2_W,
        debug=True
    )

    t = df["Time_hr"].to_numpy()

    # 1) Intracellular (IC) — per-cluster 2x2 bands
    plot_clusters_2x2_bands(
        t=t,
        mat=CbrIC_mat,
        cl=cl,
        title_prefix="Brain IC concentration (per cluster) [Option A]",
        q_lo=0.05, q_hi=0.95
    )

    # 2) Interstitial (IF) — per-cluster 2x2 bands
    plot_clusters_2x2_bands(
        t=t,
        mat=CbrIF_mat,
        cl=cl,
        title_prefix="Brain IF concentration (per cluster) [Option A]",
        q_lo=0.05, q_hi=0.95
    )

    # Optional: global system traces (venous, vascular, IF global, IC global)
    plt.figure(figsize=(10, 6))
    plt.plot(t, df["Cvpl_mg_per_L"], lw=2, label="Venous plasma (Cvpl)")
    plt.plot(t, df["CbrV_mg_per_L"], lw=2, label="Brain vascular (CbrV)")
    plt.plot(t, df["CbrIF_global_mg_per_L"], lw=2, label="Brain IF global")
    plt.plot(t, df["CbrIC_global_mg_per_L"], lw=2, label="Brain IC global")
    plt.xlabel("Time (hr)")
    plt.ylabel("mg/L")
    plt.title("AZD system-level traces [Option A]")
    plt.grid(True, ls="--", alpha=0.5)
    plt.legend()
    plt.tight_layout()
    plt.show()

    # Excel export (subset of cells per cluster)
    OUT_XLSX = "azd1775_brain_outputs.xlsx"
    export_azd_excel(
        out_xlsx=OUT_XLSX,
        sol=sol,
        df=df,
        CbrIF_mat=CbrIF_mat,
        CbrIC_mat=CbrIC_mat,
        cl=cl,
        N_tot=N_tot,
        n_keep_per_cluster=2500,
        seed_save=123
    )
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    